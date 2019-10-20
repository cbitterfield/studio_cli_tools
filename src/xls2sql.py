#!/usr/bin/env python3
# encoding: utf-8
'''
xls2sql -- shortdesc

xls2sql is a description

It defines classes_and_methods

@author:     colin


@copyright:  2019 organization_name. All rights reserved.

@license:    license

@contact:    user_email
@deffield    updated: Updated
'''

import sys
import os
import calendar
import time
import mysql.connector

# Set a timestamp for whatever we need a uniquew number for
current_time = calendar.timegm(time.gmtime())

from argparse import ArgumentParser
from argparse import RawDescriptionHelpFormatter

__all__ = []
__version__ = 0.1
__date__ = '2019-08-04'
__updated__ = '2019-08-04'

DEBUG = False
TESTRUN = False
NOEXEC = False

class CLIError(Exception):
    '''Generic exception to raise and log different fatal errors.'''
    def __init__(self, msg):
        super(CLIError).__init__(type(self))
        self.msg = "E: %s" % msg
    def __str__(self):
        return self.msg
    def __unicode__(self):
        return self.msg

def main(argv=None): # IGNORE:C0111
    '''Command line options.'''

    if argv is None:
        argv = sys.argv
    else:
        sys.argv.extend(argv)

    program_name = os.path.basename(sys.argv[0])
    program_version = "v%s" % __version__
    program_build_date = str(__updated__)
    program_version_message = '%%(prog)s %s (%s)' % (program_version, program_build_date)
    program_shortdesc = __import__('__main__').__doc__.split("\n")[1]
    program_license = '''%s

  Created by user_name on %s.
  Copyright 2019 organization_name. All rights reserved.

  Licensed under the GNU Public License 3.0
  https://www.gnu.org/licenses/gpl-3.0.en.html

  Distributed on an "AS IS" basis without warranties
  or conditions of any kind, either express or implied.

USAGE
''' % (program_shortdesc, str(__date__))

    try:
        # Setup argument parser
        parser = ArgumentParser(description=program_license, formatter_class=RawDescriptionHelpFormatter)
        parser.add_argument("-l","--log", action="store",default="console", help="logfile for output, if none is selected then use console" )
        parser.add_argument('-V', '--version', action='version', version=program_version_message)
        parser.add_argument("-i", "--input", dest="input", action="store", help="set input filename")
        parser.add_argument("-d", "--debug", dest="debug", action="store", default="INFO", help="set the debug level [INFO|ERROR|WARN|DEBUG]")
        parser.add_argument("-n", "--noexec", dest="noexec", action="store_true", help="Do not make any changes to data or os layer, just show what you would do" )
        parser.add_argument("-u", "--user", dest="db_user", action="store", help="Username of MySQL server")
        parser.add_argument("-p", "--password", dest="db_pass", action="store", help="Username of MySQL server")
        parser.add_argument("-H", "--hostname", dest="db_hostname", action="store", help="Username of MySQL server")
        parser.add_argument("-P", "--port", dest="db_port", action="store", default="3306",help="Username of MySQL server")
        parser.add_argument("-D", "--database", dest="db_schema", action="store", default="mysql",help="Database to use in  MySQL server")
        parser.add_argument("-t","--table-name",dest="db_table",action="store",default="import"+str(current_time), help="Table to receive import data")
        parser.add_argument("-s","--sheet",dest="db_sheet",action="store",default="0",help="Worksheet to extract data from")
        parser.add_argument("-r","--range",dest="db_range",action="store",help="Worksheet Range")
        
        # Process arguments
        args = parser.parse_args()
        #print(args)

        
        logfile = args.log
        debug = args.debug
        inputfile = args.input
        noexec = args.noexec
        db_user = args.db_user
        db_pass = args.db_pass
        db_host = args.db_hostname
        db_port = args.db_port
        db_table = args.db_table
        db_sheet = args.db_sheet
        db_range = args.db_range
        db_schema = args.db_schema


        if DEBUG:
            print("Debug Value: {}".format(debug))
            
        if noexec:
            print("NOEXEC=True; No data or files will be changed")

        #===============================================================================
        # Setup  Logging
        #===============================================================================
        import logging
        import logging.config 
        import logging.handlers
        

        # Set Formatting
        LOG_FORMAT = '%(asctime)s:%(name)s:%(levelname)s:%(message)s'
        LOG_DATE = '%m/%d/%Y %I:%M:%S %p'
        LOG_STYLE = style='%'

        if not 'Linux' in os.uname()[0]:
            LOG_SOCKET = '/var/run/syslog'
        else:
            LOG_SOCKET = '/dev/log'

        # create logger
        # set name as desired if not in a module
        logger = logging.getLogger('__name__')
        logger.setLevel(logging.DEBUG)

        # create handlers for Console and set level
        CONSOLE = logging.StreamHandler()
        CONSOLE.setLevel(logging.DEBUG)

        #create handlers for Syslog and set level
        SYSLOG = logging.handlers.SysLogHandler(address=LOG_SOCKET, facility='local0')
        SYSLOG.setLevel(logging.INFO)

        #create handler for FILENAME and set level
        LOG_FILE = logging.FileHandler('/tmp/logger.log',mode='a', encoding=None, delay=False)
        LOG_FILE.setLevel(logging.INFO)
        # create formatter
        formatter = logging.Formatter(LOG_FORMAT)

        # add formatter(s) to handlers
        CONSOLE.setFormatter(formatter)
        SYSLOG.setFormatter(formatter)
        LOG_FILE.setFormatter(formatter)

        # add handlers to logger
        logger.addHandler(CONSOLE)
        logger.addHandler(SYSLOG)
        logger.addHandler(LOG_FILE)
         
        
        
        #===========================================================================
        # Main program starts here
        #===========================================================================
        logger.info('{0} starting with debug level of {1}'.format(os.path.basename(__file__),debug))        
        logger.info('Connecting to {0} as {1} using schema {2}'.format(db_host,db_user,db_schema))
        logger.info('Importing {0} range {1} from  {2}'.format(db_sheet,db_range,inputfile))
        
        
        # MySQL Connector
        
        config = {
              'user': db_user,
              'password': db_pass,
              'host': db_host,
              'database': db_schema,
              'port' :db_port,
              'raise_on_warnings': False,
              'charset': 'utf8'
                }
        try:
            cnx = mysql.connector.connect(**config)
            if cnx.is_connected():
                logger.info('Connection established')
        except:
            logger.error('Connection failed')
            return 2
        
        cursor = cnx.cursor()
        
        
        
        #Add test see if the table exists and to add a ts if it does
        
        
        # Import `load_workbook` module from `openpyxl`
        from openpyxl import load_workbook
        
        # Load in the workbook
        wb = load_workbook('./test.xlsx')
        
        # Get sheet names
        print(wb.get_sheet_names())

        #=======================================================================
        # Main program ends here
        #=======================================================================

        return 0
    except KeyboardInterrupt:
        ### handle keyboard interrupt ###
        return 0
    except Exception as e:
        if DEBUG or TESTRUN:
            raise(e)
        indent = len(program_name) * " "
        sys.stderr.write(program_name + ": " + repr(e) + "\n")
        sys.stderr.write(indent + "  for help use --help")
        return 2
    
    
    
    
    

if __name__ == "__main__":
    if DEBUG:
        sys.argv.append("-d INFO")
        
    if TESTRUN:
        import doctest
        doctest.testmod()
    if NOEXEC:
        print("No exec")
        sys.argv.append("--noexec")
        
    
    



   
        
    sys.exit(main())